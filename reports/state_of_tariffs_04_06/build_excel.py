"""
Build data_download.xlsx for State of Tariffs: April 6, 2026.

Reads model output CSVs and creates formatted Excel workbook with all
tables and figure data.

Usage: python reports/state_of_tariffs_04_06/build_excel.py
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TEMP_DIR = os.path.join(REPO, 'output', '2026-04-06', 'results')
PERM_DIR = os.path.join(REPO, 'output', '2026-04-06_s122perm', 'results')
EXAMPLE_XLSX = os.path.join(REPO, 'reports', 'state_of_tariffs_example', 'data_download.xlsx')
OUT_PATH = os.path.join(REPO, 'reports', 'state_of_tariffs_04_06', 'data_download.xlsx')

# ---- Helpers ---------------------------------------------------------------

def read_csv_dict(path):
    """Read CSV into list of dicts."""
    with open(path, newline='') as f:
        return list(csv.DictReader(f))

def read_key_results(path):
    """Read key_results.csv into {metric: value} dict."""
    rows = read_csv_dict(path)
    return {r['metric']: float(r['value']) for r in rows}

# Styles
BOLD = Font(bold=True)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

# ---- Load data -------------------------------------------------------------

temp = read_key_results(os.path.join(TEMP_DIR, 'key_results.csv'))
perm = read_key_results(os.path.join(PERM_DIR, 'key_results.csv'))

temp_etrs = read_csv_dict(os.path.join(TEMP_DIR, 'goods_weighted_etrs.csv'))
perm_etrs = read_csv_dict(os.path.join(PERM_DIR, 'goods_weighted_etrs.csv'))

temp_rev = read_csv_dict(os.path.join(TEMP_DIR, 'revenue_by_year.csv'))
perm_rev = read_csv_dict(os.path.join(PERM_DIR, 'revenue_by_year.csv'))

temp_dyn = read_csv_dict(os.path.join(TEMP_DIR, 'dynamic_revenue_by_year.csv'))
perm_dyn = read_csv_dict(os.path.join(PERM_DIR, 'dynamic_revenue_by_year.csv'))

temp_macro = read_csv_dict(os.path.join(TEMP_DIR, 'macro_quarterly.csv'))
perm_macro = read_csv_dict(os.path.join(PERM_DIR, 'macro_quarterly.csv'))

temp_sectors = read_csv_dict(os.path.join(TEMP_DIR, 'sector_effects.csv'))
perm_sectors = read_csv_dict(os.path.join(PERM_DIR, 'sector_effects.csv'))

temp_foreign = read_csv_dict(os.path.join(TEMP_DIR, 'foreign_gdp.csv'))
perm_foreign = read_csv_dict(os.path.join(PERM_DIR, 'foreign_gdp.csv'))

temp_dist = read_csv_dict(os.path.join(TEMP_DIR, 'distribution.csv'))
perm_dist = read_csv_dict(os.path.join(PERM_DIR, 'distribution.csv'))

temp_dist_ps = read_csv_dict(os.path.join(TEMP_DIR, 'distribution_postsub.csv'))
perm_dist_ps = read_csv_dict(os.path.join(PERM_DIR, 'distribution_postsub.csv'))

temp_pce = read_csv_dict(os.path.join(TEMP_DIR, 'pce_major_category_prices.csv'))
perm_pce = read_csv_dict(os.path.join(PERM_DIR, 'pce_major_category_prices.csv'))

# ---- Create workbook -------------------------------------------------------

wb = Workbook()

# ==== Data TOC ==============================================================
ws = wb.active
ws.title = 'Data TOC'
ws['A1'] = 'State of U.S. Tariffs: April 6, 2026'
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = 'April 2026'
ws['A3'] = 'The Budget Lab at Yale'
ws['A5'] = 'Tables and Figures'
ws['A5'].font = BOLD
toc_items = [
    ('T1', 'Table 1. Summary Economic & Fiscal Effects'),
    ('F1', 'Figure 1. US Average Effective Tariff Rate Since 1790'),
    ('F2', 'Figure 2. Tariff Impact Tracker: Daily Effective Tariff Rate by Authority'),
    ('T2', 'Table 2. Average Effective US Tariff Rate'),
    ('F3', 'Figure 3. US Real GDP Level Effects'),
    ('F4', 'Figure 4. Change in Long-Run Real US GDP by Sector'),
    ('F5', 'Figure 5. Long-Run Change in Real GDP Level'),
    ('T3', 'Table 3. Estimated Revenue Effects'),
    ('F6', 'Figure 6. Distributional Effects'),
    ('F7', 'Figure 7. Consumer Price Effects by PCE Spending Category'),
]
for i, (sheet, desc) in enumerate(toc_items):
    ws.cell(row=6 + i, column=1, value=sheet)
    ws.cell(row=6 + i, column=2, value=desc)
auto_width(ws)

# ==== T1: Summary table =====================================================
ws = wb.create_sheet('T1')
ws['A1'] = 'Table 1. Summary Economic & Fiscal Effects of Trump Administration Tariffs'
ws['A1'].font = Font(bold=True, size=11)
ws['A2'] = 'Source: Congressional Budget Office, S&P Global, GTAP v7 [Corong et al (2017)], GTAP-RD, The Budget Lab analysis.'
ws['A2'].font = Font(italic=True, size=9)

# Headers
ws['B5'] = 'Section 122\nExpires in 150 Days'
ws['C5'] = 'Section 122\nExtended After 150 Days'
for c in ['B', 'C']:
    cell = ws[f'{c}5']
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

row = 7
ws.cell(row=row, column=1, value='Effective Tariff Rates at the End of 2026').font = BOLD
row += 1
ws.cell(row=row, column=1, value='Overall, Pre-Substitution')
ws.cell(row=row, column=2, value=temp['pre_sub_all_in_etr'] / 100)
ws.cell(row=row, column=3, value=perm['pre_sub_all_in_etr'] / 100)
ws[f'B{row}'].number_format = '0.0%'
ws[f'C{row}'].number_format = '0.0%'
row += 1
ws.cell(row=row, column=1, value='Overall, Post-Substitution')
ws.cell(row=row, column=2, value=temp['pe_postsub_all_in_etr'] / 100)
ws.cell(row=row, column=3, value=perm['pe_postsub_all_in_etr'] / 100)
ws[f'B{row}'].number_format = '0.0%'
ws[f'C{row}'].number_format = '0.0%'

row += 1
ws.cell(row=row, column=1, value='Fiscal').font = BOLD
row += 1
ws.cell(row=row, column=1, value='Conventional Revenue, 2026-2035 (Trillions)')
ws.cell(row=row, column=2, value=temp['conventional_revenue_10yr'] / 1000)
ws.cell(row=row, column=3, value=perm['conventional_revenue_10yr'] / 1000)
ws[f'B{row}'].number_format = '0.000'
ws[f'C{row}'].number_format = '0.000'
row += 1
ws.cell(row=row, column=1, value='Dynamic Revenue, 2026-2035 (Trillions)')
ws.cell(row=row, column=2, value=temp['dynamic_revenue_10yr'] / 1000)
ws.cell(row=row, column=3, value=perm['dynamic_revenue_10yr'] / 1000)
ws[f'B{row}'].number_format = '0.000'
ws[f'C{row}'].number_format = '0.000'

row += 1
ws.cell(row=row, column=1, value='Prices in the Medium Run').font = BOLD
row += 1
ws.cell(row=row, column=1, value='Percent Change in PCE Price Level, pre-substitution')
ws.cell(row=row, column=2, value=temp['pre_sub_price_increase'] / 100)
ws.cell(row=row, column=3, value=perm['pre_sub_price_increase'] / 100)
ws[f'B{row}'].number_format = '0.00%'
ws[f'C{row}'].number_format = '0.00%'
row += 1
ws.cell(row=row, column=1, value='Percent Change in PCE Price Level, post-substitution')
ws.cell(row=row, column=2, value=temp['pe_postsub_price_increase'] / 100)
ws.cell(row=row, column=3, value=perm['pe_postsub_price_increase'] / 100)
ws[f'B{row}'].number_format = '0.00%'
ws[f'C{row}'].number_format = '0.00%'
row += 1
ws.cell(row=row, column=1, value='Average Household Real Income Loss, Pre-Substitution (2025$)')
ws.cell(row=row, column=2, value=temp['pre_sub_per_hh_cost'])
ws.cell(row=row, column=3, value=perm['pre_sub_per_hh_cost'])
ws[f'B{row}'].number_format = '#,##0'
ws[f'C{row}'].number_format = '#,##0'
row += 1
ws.cell(row=row, column=1, value='Average Household Real Income Loss, Post-Substitution (2025$)')
ws.cell(row=row, column=2, value=temp['post_sub_per_hh_cost'])
ws.cell(row=row, column=3, value=perm['post_sub_per_hh_cost'])
ws[f'B{row}'].number_format = '#,##0'
ws[f'C{row}'].number_format = '#,##0'

row += 1
ws.cell(row=row, column=1, value='Output and Employment').font = BOLD
row += 1
ws.cell(row=row, column=1, value='Percentage Point Change in Q4-Q4 GDP Growth, 2026')
ws.cell(row=row, column=2, value=temp['gdp_2026_q4q4'])
ws.cell(row=row, column=3, value=perm['gdp_2026_q4q4'])
ws[f'B{row}'].number_format = '0.00'
ws[f'C{row}'].number_format = '0.00'
row += 1
# Long-run GDP from sector_effects overall_gdp
temp_lr_gdp = [r for r in temp_sectors if r['sector'] == 'overall_gdp'][0]
perm_lr_gdp = [r for r in perm_sectors if r['sector'] == 'overall_gdp'][0]
ws.cell(row=row, column=1, value='Percent change in long-run GDP')
ws.cell(row=row, column=2, value=float(temp_lr_gdp['output_change_pct']) / 100)
ws.cell(row=row, column=3, value=float(perm_lr_gdp['output_change_pct']) / 100)
ws[f'B{row}'].number_format = '0.000%'
ws[f'C{row}'].number_format = '0.000%'
row += 1
ws.cell(row=row, column=1, value='Percentage Point Change in the Unemployment Rate, End of 2026')
ws.cell(row=row, column=2, value=temp['urate_2026_q4'])
ws.cell(row=row, column=3, value=perm['urate_2026_q4'])
ws[f'B{row}'].number_format = '0.000'
ws[f'C{row}'].number_format = '0.000'

ws.column_dimensions['A'].width = 55
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 22

# ==== F1: Historical ETR ====================================================
ws = wb.create_sheet('F1')
ws['A1'] = 'Figure 1. US Average Effective Tariff Rate Since 1790'
ws['A1'].font = Font(bold=True, size=11)
ws['A2'] = 'Subtitle: Customs duty revenue as a percentage of goods imports'
ws['A3'] = 'Source: Historical Statistics of the United States Ea424-434, Monthly Treasury Statement, Bureau of Economic Analysis, The Budget Lab analysis.'
ws['A3'].font = Font(italic=True, size=9)

# Extract historical data from example xlsx
from openpyxl import load_workbook
example_wb = load_workbook(EXAMPLE_XLSX, data_only=True)
example_f1 = example_wb['F1']

# Headers
ws['A5'] = 'Year'
ws['B5'] = 'Effective Tariff Rate'
ws['C5'] = 'Projected Post-Substitution Rate'
ws['D5'] = 'Current Post-Substitution Rate'
ws['E5'] = 'Projected Pre-Substitution Rate'
ws['F5'] = 'Current Pre-Substitution Rate'

new_postsub = temp['pe_postsub_all_in_etr']
new_presub = temp['pre_sub_all_in_etr']

for src_row in range(6, example_f1.max_row + 1):
    year_val = example_f1.cell(row=src_row, column=1).value
    etr_val = example_f1.cell(row=src_row, column=2).value

    if year_val is None:
        continue

    dest_row = src_row
    ws.cell(row=dest_row, column=1, value=year_val)
    if etr_val is not None:
        ws.cell(row=dest_row, column=2, value=etr_val)

    # Horizontal reference lines
    ws.cell(row=dest_row, column=3, value=new_postsub)
    ws.cell(row=dest_row, column=5, value=new_presub)

    if year_val == 2024:
        ws.cell(row=dest_row, column=4, value=new_postsub)
        ws.cell(row=dest_row, column=6, value=new_presub)
    elif year_val == 2025:
        # 2025: observed ETR and projected rates
        ws.cell(row=dest_row, column=3, value=7.7303)
        ws.cell(row=dest_row, column=4, value=7.7303)
        ws.cell(row=dest_row, column=5, value=7.7303)
        ws.cell(row=dest_row, column=6, value=7.7303)
    elif year_val == 2026:
        ws.cell(row=dest_row, column=3, value=new_postsub)
        ws.cell(row=dest_row, column=5, value=new_presub)

example_wb.close()
auto_width(ws)

# ==== F2: Tariff Impact Tracker (PLACEHOLDER) ================================
ws = wb.create_sheet('F2')
ws['A1'] = 'Figure 2. Tariff Impact Tracker: Daily Effective Tariff Rate by Authority'
ws['A1'].font = Font(bold=True, size=11)
ws['A2'] = 'Subtitle: Percent'
ws['A3'] = 'Source: The Budget Lab analysis.'
ws['A3'].font = Font(italic=True, size=9)
ws['A5'] = 'PLACEHOLDER - User will fill in daily ETR data by authority'
ws['A5'].font = Font(bold=True, color='FF0000')

# ==== T2: ETR by country ====================================================
ws = wb.create_sheet('T2')
ws['A1'] = 'Table 2. Average Effective US Tariff Rate at the End of 2026, Trump Administration Tariffs'
ws['A1'].font = Font(bold=True, size=11)
ws['A2'] = 'Subtitle: Import-weighted tariff rate levels, pre- and post-substitution'
ws['A3'] = 'Source: GTAP v7, The Budget Lab analysis.'
ws['A3'].font = Font(italic=True, size=9)

# --- Section 122 Expires ---
ws['A5'] = 'Section 122 Expires'
ws['A5'].font = BOLD

headers = ['', 'Avg Effective Tariff Rate\nPre-Sub', 'Avg Effective Tariff Rate\nPost-Sub',
           'Import Share\nPre-Sub', 'Import Share\nPost-Sub',
           'Contribution\nPre-Sub', 'Contribution\nPost-Sub']
for j, h in enumerate(headers):
    cell = ws.cell(row=6, column=j+1, value=h)
    if j > 0:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

country_display = {
    'chn': 'China', 'ca': 'Canada', 'mx': 'Mexico',
    'eu': 'EU', 'jp': 'Japan', 'uk': 'UK',
    'fta': 'FTA Partners', 'row': 'Rest of World', 'all': 'Total',
}
country_order = ['chn', 'ca', 'mx', 'eu', 'jp', 'uk', 'fta', 'row', 'all']

def write_etr_section(ws, start_row, etrs_list):
    etrs_dict = {r['country_code']: r for r in etrs_list}
    total_presub = float(etrs_dict['all']['presub_imports'])
    total_postsub = float(etrs_dict['all']['pe_postsub_imports'])

    row = start_row
    for code in country_order:
        r = etrs_dict[code]
        name = country_display[code]
        presub_level = float(r['presub_level'])
        postsub_level = float(r['pe_postsub_level'])
        presub_imp = float(r['presub_imports'])
        postsub_imp = float(r['pe_postsub_imports'])
        presub_share = presub_imp / total_presub
        postsub_share = postsub_imp / total_postsub
        presub_contrib = presub_level * presub_share
        postsub_contrib = postsub_level * postsub_share

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=presub_level)
        ws.cell(row=row, column=3, value=postsub_level)
        ws.cell(row=row, column=4, value=presub_share)
        ws.cell(row=row, column=5, value=postsub_share)
        ws.cell(row=row, column=6, value=presub_contrib)
        ws.cell(row=row, column=7, value=postsub_contrib)

        for col in range(2, 8):
            ws.cell(row=row, column=col).number_format = '0.00'
        for col in [4, 5]:
            ws.cell(row=row, column=col).number_format = '0.00%'

        if name == 'Total':
            ws.cell(row=row, column=1).font = BOLD
        row += 1
    return row

r = 7
r = write_etr_section(ws, r, temp_etrs)

# --- Section 122 Extended ---
r += 1
ws.cell(row=r, column=1, value='Section 122 Extended').font = BOLD
r += 1
for j, h in enumerate(headers):
    cell = ws.cell(row=r, column=j+1, value=h)
    if j > 0:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
r += 1
write_etr_section(ws, r, perm_etrs)
auto_width(ws)

# ==== F3: GDP Level Effects (quarterly) ======================================
ws = wb.create_sheet('F3')
ws['A1'] = 'Figure 3. US Real GDP Level Effects of Trump Administration Tariffs'
ws['A1'].font = Font(bold=True, size=11)
ws['A2'] = 'Subtitle: Percentage point change against baseline'
ws['A3'] = 'Source: S&P Global, GTAP v7, The Budget Lab analysis.'
ws['A3'].font = Font(italic=True, size=9)

ws['A5'] = 'Date'
ws['B5'] = 'Section 122 Expires'
ws['C5'] = 'Section 122 Extended'
style_header_row(ws, 5, 3)

for i, (tr, pr) in enumerate(zip(temp_macro, perm_macro)):
    row = 6 + i
    yr = int(tr['year'])
    qtr = int(tr['quarter'])
    temp_gdp_delta = (float(tr['gdp_tariff']) / float(tr['gdp_baseline']) - 1) * 100
    perm_gdp_delta = (float(pr['gdp_tariff']) / float(pr['gdp_baseline']) - 1) * 100

    ws.cell(row=row, column=1, value=f'{yr}Q{qtr}')
    ws.cell(row=row, column=2, value=temp_gdp_delta)
    ws.cell(row=row, column=3, value=perm_gdp_delta)
auto_width(ws)

# ==== F4: Sectoral GDP ======================================================
ws = wb.create_sheet('F4')
ws['A1'] = 'Figure 4. Change in Long-Run Real US GDP by Sector from Trump Administration Tariffs'
ws['A1'].font = Font(bold=True, size=11)
ws['A2'] = 'Subtitle: Percentage points'
ws['A3'] = 'Notes: Real value added by sector.'
ws['A4'] = 'Source: GTAP v7, The Budget Lab analysis.'
ws['A4'].font = Font(italic=True, size=9)

ws['B6'] = 'Section 122 Expires'
ws['C6'] = 'Section 122 Extended'
style_header_row(ws, 6, 3)

sector_display = {
    'agriculture': 'Agriculture',
    'mining': 'Mining & Extraction',
    'manufacturing': 'Total Manufacturing',
    'durable': 'Durable Manufacturing',
    'advanced': 'Advanced Manufacturing',
    'nondurable': 'Nondurable Manufacturing',
    'utilities': 'Utilities',
    'construction': 'Construction',
    'services': 'Services',
    'overall_gdp': 'Overall Real GDP',
}
sector_order = ['agriculture', 'mining', 'manufacturing', 'durable', 'advanced',
                'nondurable', 'utilities', 'construction', 'services', 'overall_gdp']

temp_sec_dict = {r['sector']: float(r['output_change_pct']) for r in temp_sectors}
perm_sec_dict = {r['sector']: float(r['output_change_pct']) for r in perm_sectors}

for i, sec in enumerate(sector_order):
    row = 7 + i
    ws.cell(row=row, column=1, value=sector_display[sec])
    ws.cell(row=row, column=2, value=temp_sec_dict[sec])
    ws.cell(row=row, column=3, value=perm_sec_dict[sec])
auto_width(ws)

# ==== F5: Foreign GDP ========================================================
ws = wb.create_sheet('F5')
ws['A1'] = 'Figure 5. Long-Run Change in Real GDP Level from Trump Administration Tariffs'
ws['A1'].font = Font(bold=True, size=11)
ws['A2'] = 'Subtitle: Percentage point change'
ws['A3'] = 'Notes: FTROW = countries with a comprehensive free trade agreement with the US. ROW = all other countries.'
ws['A4'] = 'Source: GTAP v7 [Corong et al (2017)], The Budget Lab analysis.'
ws['A4'].font = Font(italic=True, size=9)

ws['B6'] = 'Section 122 Expires'
ws['C6'] = 'Section 122 Extended'
style_header_row(ws, 6, 3)

region_display = {
    'usa': 'USA', 'china': 'China', 'row': 'ROW', 'canada': 'Canada',
    'mexico': 'Mexico', 'fta': 'FTROW', 'japan': 'Japan', 'eu': 'EU',
    'uk': 'UK', 'world': 'World Total', 'world_ex_usa': 'World ex USA'
}
region_order = ['usa', 'china', 'row', 'canada', 'mexico', 'fta', 'japan', 'eu', 'uk', 'world', 'world_ex_usa']

temp_for_dict = {r['region']: float(r['gdp_change_pct']) for r in temp_foreign}
perm_for_dict = {r['region']: float(r['gdp_change_pct']) for r in perm_foreign}

for i, reg in enumerate(region_order):
    row = 7 + i
    ws.cell(row=row, column=1, value=region_display[reg])
    ws.cell(row=row, column=2, value=temp_for_dict[reg])
    ws.cell(row=row, column=3, value=perm_for_dict[reg])
auto_width(ws)

# ==== T3: Revenue by Year ===================================================
ws = wb.create_sheet('T3')
ws['A1'] = 'Table 3. Estimated Revenue Effects of Trump Administration Tariffs'
ws['A1'].font = Font(bold=True, size=11)
ws['A3'] = 'Source: Congressional Budget Office, GTAP v7 [Corong et al (2017)], The Budget Lab analysis.'
ws['A3'].font = Font(italic=True, size=9)

years = [str(y) for y in range(2026, 2036)]

# --- Section 122 Expires ---
ws['A5'] = 'Section 122 Expires'
ws['A5'].font = BOLD

for j, y in enumerate(years):
    ws.cell(row=6, column=j+2, value=int(y))
ws.cell(row=6, column=len(years)+2, value='2026-35')
style_header_row(ws, 6, len(years) + 2)

ws.cell(row=7, column=1, value='Conventional')
conv_total = 0
for j, y in enumerate(years):
    rev_row = [r for r in temp_rev if r['fiscal_year'] == y]
    if rev_row:
        val = float(rev_row[0]['net_revenue'])
        ws.cell(row=7, column=j+2, value=round(val, 1))
        conv_total += val
ws.cell(row=7, column=len(years)+2, value=round(conv_total, 1))

ws.cell(row=8, column=1, value='Dynamic')
dyn_total = 0
for j, y in enumerate(years):
    dyn_row = [r for r in temp_dyn if r['fiscal_year'] == y]
    if dyn_row:
        val = float(dyn_row[0]['dynamic_revenue'])
        ws.cell(row=8, column=j+2, value=round(val, 1))
        dyn_total += val
ws.cell(row=8, column=len(years)+2, value=round(dyn_total, 1))

ws.cell(row=9, column=1, value='Dynamic effect')
eff_total = 0
for j, y in enumerate(years):
    dyn_row = [r for r in temp_dyn if r['fiscal_year'] == y]
    if dyn_row:
        val = float(dyn_row[0]['dynamic_effect'])
        ws.cell(row=9, column=j+2, value=round(val, 1))
        eff_total += val
ws.cell(row=9, column=len(years)+2, value=round(eff_total, 1))

# --- Section 122 Extended ---
ws['A11'] = 'Section 122 Extended'
ws['A11'].font = BOLD
for j, y in enumerate(years):
    ws.cell(row=12, column=j+2, value=int(y))
ws.cell(row=12, column=len(years)+2, value='2026-35')
style_header_row(ws, 12, len(years) + 2)

ws.cell(row=13, column=1, value='Conventional')
conv_total = 0
for j, y in enumerate(years):
    rev_row = [r for r in perm_rev if r['fiscal_year'] == y]
    if rev_row:
        val = float(rev_row[0]['net_revenue'])
        ws.cell(row=13, column=j+2, value=round(val, 1))
        conv_total += val
ws.cell(row=13, column=len(years)+2, value=round(conv_total, 1))

ws.cell(row=14, column=1, value='Dynamic')
dyn_total = 0
for j, y in enumerate(years):
    dyn_row = [r for r in perm_dyn if r['fiscal_year'] == y]
    if dyn_row:
        val = float(dyn_row[0]['dynamic_revenue'])
        ws.cell(row=14, column=j+2, value=round(val, 1))
        dyn_total += val
ws.cell(row=14, column=len(years)+2, value=round(dyn_total, 1))

ws.cell(row=15, column=1, value='Dynamic effect')
eff_total = 0
for j, y in enumerate(years):
    dyn_row = [r for r in perm_dyn if r['fiscal_year'] == y]
    if dyn_row:
        val = float(dyn_row[0]['dynamic_effect'])
        ws.cell(row=15, column=j+2, value=round(val, 1))
        eff_total += val
ws.cell(row=15, column=len(years)+2, value=round(eff_total, 1))

ws.column_dimensions['A'].width = 18
for j in range(2, len(years) + 3):
    ws.column_dimensions[chr(64 + j) if j <= 26 else 'A' + chr(64 + j - 26)].width = 10

# ==== F6: Distribution ======================================================
ws = wb.create_sheet('F6')
ws['A1'] = 'Figure 6. Distributional Effects of Trump Administration Tariffs'
ws['A1'].font = Font(bold=True, size=11)
ws['A2'] = 'Subtitle: By household income decile'
ws['A3'] = 'Source: GTAP v7, Census, BLS, BEA, The Budget Lab analysis.'
ws['A3'].font = Font(italic=True, size=9)

# Section 122 Expires - % of ATI (pre-sub)
ws['A5'] = 'Section 122 Expires'
ws['A5'].font = BOLD
ws['B6'] = 'Decile'
for i in range(1, 11):
    ws.cell(row=6, column=i+1, value=i)
style_header_row(ws, 6, 11)

ws.cell(row=7, column=1, value='% of ATI')
for i, d in enumerate(sorted(temp_dist, key=lambda x: int(x['decile']))):
    ws.cell(row=7, column=i+2, value=float(d['pct_of_income']))

ws.cell(row=9, column=1, value='2025$')
for i in range(1, 11):
    ws.cell(row=9, column=i+1, value=i)

ws.cell(row=10, column=1, value='Cost per HH')
for i, d in enumerate(sorted(temp_dist, key=lambda x: int(x['decile']))):
    ws.cell(row=10, column=i+2, value=float(d['cost_per_hh']))

# Section 122 Extended
r = 12
ws.cell(row=r, column=1, value='Section 122 Extended').font = BOLD
r += 1
ws.cell(row=r, column=1, value='Decile')
for i in range(1, 11):
    ws.cell(row=r, column=i+1, value=i)
style_header_row(ws, r, 11)

r += 1
ws.cell(row=r, column=1, value='% of ATI')
for i, d in enumerate(sorted(perm_dist, key=lambda x: int(x['decile']))):
    ws.cell(row=r, column=i+2, value=float(d['pct_of_income']))

r += 2
ws.cell(row=r, column=1, value='2025$')
for i in range(1, 11):
    ws.cell(row=r, column=i+1, value=i)

r += 1
ws.cell(row=r, column=1, value='Cost per HH')
for i, d in enumerate(sorted(perm_dist, key=lambda x: int(x['decile']))):
    ws.cell(row=r, column=i+2, value=float(d['cost_per_hh']))
auto_width(ws)

# ==== F7: PCE Major Category Prices ==========================================
ws = wb.create_sheet('F7')
ws['A1'] = 'Figure 7. Consumer Price Effects by PCE Spending Category'
ws['A1'].font = Font(bold=True, size=11)
ws['A2'] = 'Subtitle: Percentage point change in consumer prices'
ws['A3'] = 'Source: BEA, GTAP v7, The Budget Lab analysis.'
ws['A3'].font = Font(italic=True, size=9)

# Section 122 Expires
ws['A5'] = 'Section 122 Expires'
ws['A5'].font = BOLD
headers = ['Major Category', 'Top Level', 'Pre-Sub (%)', 'PE Post-Sub (%)', 'GE (%)', 'PCE Share (%)']
for j, h in enumerate(headers):
    ws.cell(row=6, column=j+1, value=h)
style_header_row(ws, 6, len(headers))

for i, r_data in enumerate(temp_pce):
    row = 7 + i
    ws.cell(row=row, column=1, value=r_data['major_category'])
    ws.cell(row=row, column=2, value=r_data['top_level'])
    ws.cell(row=row, column=3, value=float(r_data['pre_sub']))
    ws.cell(row=row, column=4, value=float(r_data['pe_postsub']))
    ws.cell(row=row, column=5, value=float(r_data['ge']))
    ws.cell(row=row, column=6, value=float(r_data['pce_share']))

# Section 122 Extended
r = 7 + len(temp_pce) + 2
ws.cell(row=r, column=1, value='Section 122 Extended').font = BOLD
r += 1
for j, h in enumerate(headers):
    ws.cell(row=r, column=j+1, value=h)
style_header_row(ws, r, len(headers))
r += 1

for i, r_data in enumerate(perm_pce):
    row = r + i
    ws.cell(row=row, column=1, value=r_data['major_category'])
    ws.cell(row=row, column=2, value=r_data['top_level'])
    ws.cell(row=row, column=3, value=float(r_data['pre_sub']))
    ws.cell(row=row, column=4, value=float(r_data['pe_postsub']))
    ws.cell(row=row, column=5, value=float(r_data['ge']))
    ws.cell(row=row, column=6, value=float(r_data['pce_share']))

auto_width(ws)

# ---- Save ------------------------------------------------------------------
wb.save(OUT_PATH)
print(f'Wrote {OUT_PATH}')
print(f'Sheets: {wb.sheetnames}')

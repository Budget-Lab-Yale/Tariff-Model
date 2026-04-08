"""
Refresh TBL-formatted data download with 04-06 model output.

Copies the formatted 04-02 template, overwrites only data cells,
saves as data_download.xlsx. Preserves all formatting and structure.

Usage: python reports/state_of_tariffs_04_06/refresh_numbers.py
"""

import csv
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TEMPLATE = os.path.join('C:/Users/jar335/Downloads', 'TBL-Data-State-of-Tariffs-20260402.xlsx')
OUTPUT = os.path.join(REPO, 'reports', 'state_of_tariffs_04_06', 'data_download.xlsx')
TEMP_DIR = os.path.join(REPO, 'output', '2026-04-06', 'results')
PERM_DIR = os.path.join(REPO, 'output', '2026-04-06_s122perm', 'results')


def read_csv_dict(path):
    with open(path, newline='') as f:
        return list(csv.DictReader(f))

def read_key_results(path):
    rows = read_csv_dict(path)
    return {r['metric']: float(r['value']) for r in rows}


# ---- Load model output -----------------------------------------------------

temp = read_key_results(os.path.join(TEMP_DIR, 'key_results.csv'))
perm = read_key_results(os.path.join(PERM_DIR, 'key_results.csv'))

temp_etrs = read_csv_dict(os.path.join(TEMP_DIR, 'goods_weighted_etrs.csv'))
perm_etrs = read_csv_dict(os.path.join(PERM_DIR, 'goods_weighted_etrs.csv'))

temp_macro = read_csv_dict(os.path.join(TEMP_DIR, 'macro_quarterly.csv'))
perm_macro = read_csv_dict(os.path.join(PERM_DIR, 'macro_quarterly.csv'))

temp_sectors = read_csv_dict(os.path.join(TEMP_DIR, 'sector_effects.csv'))
perm_sectors = read_csv_dict(os.path.join(PERM_DIR, 'sector_effects.csv'))

temp_foreign = read_csv_dict(os.path.join(TEMP_DIR, 'foreign_gdp.csv'))
perm_foreign = read_csv_dict(os.path.join(PERM_DIR, 'foreign_gdp.csv'))

temp_rev = read_csv_dict(os.path.join(TEMP_DIR, 'revenue_by_year.csv'))
perm_rev = read_csv_dict(os.path.join(PERM_DIR, 'revenue_by_year.csv'))

temp_dyn = read_csv_dict(os.path.join(TEMP_DIR, 'dynamic_revenue_by_year.csv'))
perm_dyn = read_csv_dict(os.path.join(PERM_DIR, 'dynamic_revenue_by_year.csv'))

temp_dist = sorted(read_csv_dict(os.path.join(TEMP_DIR, 'distribution.csv')),
                   key=lambda x: int(x['decile']))
perm_dist = sorted(read_csv_dict(os.path.join(PERM_DIR, 'distribution.csv')),
                   key=lambda x: int(x['decile']))

temp_pce = read_csv_dict(os.path.join(TEMP_DIR, 'pce_major_category_prices.csv'))
perm_pce = read_csv_dict(os.path.join(PERM_DIR, 'pce_major_category_prices.csv'))

temp_sec_dict = {r['sector']: float(r['output_change_pct']) for r in temp_sectors}
perm_sec_dict = {r['sector']: float(r['output_change_pct']) for r in perm_sectors}

temp_for_dict = {r['region']: float(r['gdp_change_pct']) for r in temp_foreign}
perm_for_dict = {r['region']: float(r['gdp_change_pct']) for r in perm_foreign}

temp_macro_dict = {(r['year'], r['quarter']): r for r in temp_macro}
perm_macro_dict = {(r['year'], r['quarter']): r for r in perm_macro}


# ---- Copy template and open -------------------------------------------------

shutil.copy2(TEMPLATE, OUTPUT)
wb = load_workbook(OUTPUT)
print(f'Sheets: {wb.sheetnames}')


# ===========================================================================
# Data TOC: Update date
# ===========================================================================
ws = wb['Data TOC']
ws['A1'] = 'State of U.S. Tariffs: April 6, 2026'
print('  Updated: Data TOC')


# ===========================================================================
# T1: Summary
# ===========================================================================
ws = wb['T1']

# Row 7: Overall ETR, Pre-Sub
ws['B7'] = temp['pre_sub_all_in_etr'] / 100
ws['C7'] = perm['pre_sub_all_in_etr'] / 100

# Row 8: Overall ETR, Post-Sub
ws['B8'] = temp['pe_postsub_all_in_etr'] / 100
ws['C8'] = perm['pe_postsub_all_in_etr'] / 100

# Row 10: Conventional Revenue (trillions)
ws['B10'] = temp['conventional_revenue_10yr'] / 1000
ws['C10'] = perm['conventional_revenue_10yr'] / 1000

# Row 11: Dynamic Revenue (trillions)
ws['B11'] = temp['dynamic_revenue_10yr'] / 1000
ws['C11'] = perm['dynamic_revenue_10yr'] / 1000

# Row 13: Price level, pre-sub
ws['B13'] = temp['pre_sub_price_increase'] / 100
ws['C13'] = perm['pre_sub_price_increase'] / 100

# Row 14: Price level, post-sub
ws['B14'] = temp['pe_postsub_price_increase'] / 100
ws['C14'] = perm['pe_postsub_price_increase'] / 100

# Row 15: HH loss, pre-sub ($)
ws['B15'] = temp['pre_sub_per_hh_cost']
ws['C15'] = perm['pre_sub_per_hh_cost']

# Row 16: HH loss, post-sub ($)
ws['B16'] = temp['post_sub_per_hh_cost']
ws['C16'] = perm['post_sub_per_hh_cost']

# Row 18: Q4-Q4 GDP Growth 2026 (pp)
ws['B18'] = temp['gdp_2026_q4q4']
ws['C18'] = perm['gdp_2026_q4q4']

# Row 19: Q4 2026 GDP level change
temp_q4 = temp_macro_dict[('2026', '4')]
perm_q4 = perm_macro_dict[('2026', '4')]
ws['B19'] = float(temp_q4['gdp_tariff']) / float(temp_q4['gdp_baseline']) - 1
ws['C19'] = float(perm_q4['gdp_tariff']) / float(perm_q4['gdp_baseline']) - 1

# Row 20: Long-run GDP
ws['B20'] = temp_sec_dict['overall_gdp'] / 100
ws['C20'] = perm_sec_dict['overall_gdp'] / 100

# Row 21: Unemployment rate change (pp)
ws['B21'] = temp['urate_2026_q4']
ws['C21'] = perm['urate_2026_q4']

print('  Updated: T1')


# ===========================================================================
# F1: Historical ETR - update reference lines
# ===========================================================================
ws = wb['F1']

new_postsub = temp['pe_postsub_all_in_etr']
new_presub = temp['pre_sub_all_in_etr']

# Columns D (4) and F (6) hold the constant reference lines
for row in range(6, 243):
    yr = ws.cell(row=row, column=1).value
    if yr is None:
        continue
    ws.cell(row=row, column=4).value = new_postsub
    ws.cell(row=row, column=6).value = new_presub

# 2024: projected lines in C and E show the 2024 observed ETR
for row in range(6, 243):
    yr = ws.cell(row=row, column=1).value
    if yr == 2024:
        obs = ws.cell(row=row, column=2).value
        ws.cell(row=row, column=3).value = obs
        ws.cell(row=row, column=5).value = obs
    elif yr == 2025:
        ws.cell(row=row, column=3).value = 7.7303
        ws.cell(row=row, column=5).value = 7.7303
    elif yr == 2026:
        ws.cell(row=row, column=3).value = new_postsub
        ws.cell(row=row, column=5).value = new_presub

print('  Updated: F1')


# ===========================================================================
# F2: Daily ETR Tracker
# ===========================================================================
ws = wb['F2']

daily_etr_file = os.path.join(TEMP_DIR, 'daily_etr_levels.csv')
if os.path.exists(daily_etr_file):
    daily_etr = read_csv_dict(daily_etr_file)

    # Pad from 2025-01-01 to first model date with baseline ETR
    from datetime import timedelta
    start = datetime(2025, 1, 1)
    first_model_date = datetime.strptime(daily_etr[0]['date'], '%Y-%m-%d')
    baseline_etr = float(daily_etr[0]['etr_level'])
    pad = []
    d = start
    while d < first_model_date:
        pad.append({'date': d.strftime('%Y-%m-%d'), 'etr_level': str(baseline_etr)})
        d += timedelta(days=1)
    daily_etr = pad + daily_etr

    # Clear existing data rows (start at row 6)
    for row in range(6, ws.max_row + 1):
        ws.cell(row=row, column=1).value = None
        ws.cell(row=row, column=2).value = None

    # Write new daily data
    for i, r in enumerate(daily_etr):
        row = 6 + i
        ws.cell(row=row, column=1).value = datetime.strptime(r['date'], '%Y-%m-%d')
        ws.cell(row=row, column=2).value = float(r['etr_level'])

    print(f'  Updated: F2 ({len(daily_etr)} days)')
else:
    print('  Skipped: F2 (daily_etr_levels.csv not found)')


# ===========================================================================
# T2: Country ETRs
# ===========================================================================
ws = wb['T2']

country_order = ['chn', 'ca', 'mx', 'eu', 'jp', 'uk', 'fta', 'row']

def write_etr_block(ws, start_row, etrs_list):
    total = [r for r in etrs_list if r['country_code'] == 'all'][0]
    total_pi = float(total['presub_imports'])
    total_psi = float(total['pe_postsub_imports'])

    etrs_by_code = {r['country_code']: r for r in etrs_list}

    for i, code in enumerate(country_order + ['all']):
        r = etrs_by_code[code]
        row = start_row + i
        pe = float(r['presub_level'])
        pse = float(r['pe_postsub_level'])
        pi = float(r['presub_imports'])
        psi = float(r['pe_postsub_imports'])
        ws.cell(row=row, column=2).value = pe
        ws.cell(row=row, column=3).value = pse
        ws.cell(row=row, column=4).value = pi / total_pi
        ws.cell(row=row, column=5).value = psi / total_psi
        ws.cell(row=row, column=6).value = pe * (pi / total_pi)
        ws.cell(row=row, column=7).value = pse * (psi / total_psi)

write_etr_block(ws, 8, temp_etrs)
write_etr_block(ws, 21, perm_etrs)

print('  Updated: T2')


# ===========================================================================
# F3: GDP Level Effects
# ===========================================================================
ws = wb['F3']

mid_month = {1: 3, 2: 6, 3: 9, 4: 12}
row = 6
for tm, pm in zip(temp_macro, perm_macro):
    yr, qtr = int(tm['year']), int(tm['quarter'])
    if yr < 2025:
        continue
    dt = datetime(yr, mid_month[qtr], 15)
    t_delta = (float(tm['gdp_tariff']) / float(tm['gdp_baseline']) - 1) * 100
    p_delta = (float(pm['gdp_tariff']) / float(pm['gdp_baseline']) - 1) * 100
    ws.cell(row=row, column=1).value = dt
    ws.cell(row=row, column=2).value = t_delta
    ws.cell(row=row, column=3).value = p_delta
    row += 1

print('  Updated: F3')


# ===========================================================================
# F4: Sector Effects
# ===========================================================================
ws = wb['F4']

sector_order = ['agriculture', 'mining', 'manufacturing', 'durable', 'advanced',
                'nondurable', 'utilities', 'construction', 'services', 'overall_gdp']

for i, sec in enumerate(sector_order):
    row = 7 + i
    ws.cell(row=row, column=2).value = temp_sec_dict[sec]
    ws.cell(row=row, column=3).value = perm_sec_dict[sec]

print('  Updated: F4')


# ===========================================================================
# F5: Foreign GDP
# ===========================================================================
ws = wb['F5']

region_order = ['usa', 'china', 'row', 'canada', 'mexico', 'fta', 'japan', 'eu',
                'uk', 'world', 'world_ex_usa']

for i, reg in enumerate(region_order):
    row = 7 + i
    ws.cell(row=row, column=2).value = temp_for_dict[reg]
    ws.cell(row=row, column=3).value = perm_for_dict[reg]

print('  Updated: F5')


# ===========================================================================
# T3: Revenue
# ===========================================================================
ws = wb['T3']

years = [str(y) for y in range(2026, 2036)]

def write_revenue_block(ws, conv_row, dyn_row, eff_row, rev_data, dyn_data):
    conv_total, dyn_total, eff_total = 0, 0, 0
    for j, y in enumerate(years):
        rev = [r for r in rev_data if r['fiscal_year'] == y]
        dyn = [r for r in dyn_data if r['fiscal_year'] == y]
        if rev:
            val = float(rev[0]['net_revenue'])
            ws.cell(row=conv_row, column=j + 2).value = val
            conv_total += val
        if dyn:
            dval = float(dyn[0]['dynamic_revenue'])
            ws.cell(row=dyn_row, column=j + 2).value = dval
            dyn_total += dval
            eval_ = float(dyn[0]['dynamic_effect'])
            ws.cell(row=eff_row, column=j + 2).value = eval_
            eff_total += eval_
    ws.cell(row=conv_row, column=12).value = conv_total
    ws.cell(row=dyn_row, column=12).value = dyn_total
    ws.cell(row=eff_row, column=12).value = eff_total

write_revenue_block(ws, 7, 8, 9, temp_rev, temp_dyn)
write_revenue_block(ws, 13, 14, 15, perm_rev, perm_dyn)

print('  Updated: T3')


# ===========================================================================
# F6: Distribution
# ===========================================================================
ws = wb['F6']

# Section 122 Expires
for i, d in enumerate(temp_dist):
    ws.cell(row=9, column=i + 2).value = -abs(float(d['pct_of_income']))
for i, d in enumerate(temp_dist):
    ws.cell(row=13, column=i + 2).value = -abs(float(d['cost_per_hh']))

# Section 122 Extended
for i, d in enumerate(perm_dist):
    ws.cell(row=19, column=i + 2).value = -abs(float(d['pct_of_income']))
for i, d in enumerate(perm_dist):
    ws.cell(row=23, column=i + 2).value = -abs(float(d['cost_per_hh']))

print('  Updated: F6')


# ===========================================================================
# F7: PCE Category Prices
# ===========================================================================
ws = wb['F7']

# Template layout: A=category, B=top_level, C=temp pre-sub, D=temp post-sub,
#                  E=perm pre-sub, F=perm post-sub, starting at row 8
perm_pce_dict = {r['major_category']: r for r in perm_pce}

for i, rd in enumerate(temp_pce):
    row = 8 + i
    ws.cell(row=row, column=1).value = rd['major_category']
    ws.cell(row=row, column=2).value = rd['top_level']
    ws.cell(row=row, column=3).value = float(rd['pre_sub'])
    ws.cell(row=row, column=4).value = float(rd['pe_postsub'])
    pr = perm_pce_dict.get(rd['major_category'])
    if pr:
        ws.cell(row=row, column=5).value = float(pr['pre_sub'])
        ws.cell(row=row, column=6).value = float(pr['pe_postsub'])

print('  Updated: F7')


# ===========================================================================
# Save
# ===========================================================================
wb.save(OUTPUT)
print(f'\nSaved: {OUTPUT}')

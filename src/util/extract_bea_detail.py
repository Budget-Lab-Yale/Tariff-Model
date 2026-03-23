# =============================================================================
# extract_bea_detail.py - Extract 2017 detail BEA I-O tables to CSVs
# =============================================================================
#
# Extracts detail-level (~400 commodities) BEA I-O tables from Excel into
# CSV files matching the same format as the existing summary-level CSVs.
#
# Sources:
#   - ImportMatrices_Before_Redefinitions_Detail.xlsx (from MAKE-USE-IMPORTS zip)
#   - IOUse_Before_Redefinitions_PRO_Detail.xlsx (from MAKE-USE-IMPORTS zip)
#   - IxC_Domestic_Detail.xlsx (from TOTAL AND DOMESTIC REQUIREMENTS zip)
#   - PCEBridge_Detail.xlsx
#
# Prerequisites:
#   pip install openpyxl
#
# Usage:
#   python src/util/extract_bea_detail.py
#
# =============================================================================

import csv
import os
import zipfile

import openpyxl


# =============================================================================
# Configuration
# =============================================================================

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
IO_DIR = os.path.join(PROJECT_ROOT, 'resources', 'io')
OUTPUT_DIR = os.path.join(IO_DIR, 'detail')
SHEET = '2017'

# Source Excel files (extracted from zips)
IMPORT_XLSX = os.path.join(OUTPUT_DIR, 'ImportMatrices_After_Redefinitions_Detail.xlsx')
USE_XLSX = os.path.join(OUTPUT_DIR, 'IOUse_After_Redefinitions_PRO_Detail.xlsx')
IXC_XLSX = os.path.join(OUTPUT_DIR, 'IxC_Domestic_Detail.xlsx')
PCE_BRIDGE_XLSX = os.path.join(IO_DIR, 'PCEBridge_Detail.xlsx')

# Summary-level crosswalk (to expand for detail)
SUMMARY_CROSSWALK = os.path.join(IO_DIR, 'gtap_bea_crosswalk.csv')


# =============================================================================
# Helpers
# =============================================================================

def extract_from_zip(zip_path, filename, dest_dir):
    """Extract a single file from a zip if not already present."""
    dest = os.path.join(dest_dir, filename)
    if os.path.exists(dest):
        print(f'  Already extracted: {filename}')
        return dest
    print(f'  Extracting {filename} from {os.path.basename(zip_path)}...')
    with zipfile.ZipFile(zip_path, 'r') as z:
        z.extract(filename, dest_dir)
    return dest


def safe_float(val):
    """Convert cell value to float, treating None/empty as 0."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def write_csv(path, header, rows):
    """Write CSV with given header and rows."""
    with open(path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)
    print(f'  Wrote {path} ({len(rows)} rows x {len(header)} cols)')


# =============================================================================
# Use table extraction (import and total)
# =============================================================================

def read_use_table(xlsx_path):
    """Read a BEA use table (import or total) from the 2017 detail sheet.

    Layout:
      Rows 5-6: headers (names/codes)
      Rows 7+: commodity data until a T-code row (T005 = Total Intermediate)
      Cols 1-2: commodity code + description
      Cols 3+: industry codes until T001 column, then final demand F-columns

    Returns:
      commodity_codes: list of str
      commodity_descs: list of str
      industry_codes: list of str
      data_matrix: list of lists (commodity x industry, float values)
      extra_rows: dict of {row_code: list of float} for T005, V*, T006, T008
      pce_col_values: dict of {commodity_code: float} for F01000 column
      commodity_full_totals: dict of {commodity_code: float} total use
        (intermediate + final demand, excluding F05000 import balancing column)
    """
    print(f'  Reading {os.path.basename(xlsx_path)} sheet {SHEET}...')
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[SHEET]

    # Read row 6 to get column codes
    row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]

    # Find T001 column (end of industry columns) and final demand columns
    t001_idx = None
    f01000_idx = None
    final_demand_indices = []
    for i, val in enumerate(row6):
        code = str(val).strip() if val else ''
        if code == 'T001':
            t001_idx = i
        elif code == 'F01000':
            f01000_idx = i
        # Final demand columns for domestic absorption:
        # Exclude F04/F04000 (exports — not domestic expenditure)
        # Exclude F05/F05000 (negative import balancing entry)
        if code.startswith('F') and not code.startswith(('F04', 'F05')):
            final_demand_indices.append(i)

    # Industry codes: col 3 (idx 2) to T001 (exclusive)
    industry_indices = list(range(2, t001_idx))
    industry_codes = []
    for i in industry_indices:
        code = str(row6[i]).strip() if row6[i] else ''
        industry_codes.append(code)

    # All use indices = industry + final demand (for full commodity totals)
    all_use_indices = industry_indices + final_demand_indices

    # Read all data rows (row 7 onwards)
    commodity_codes = []
    commodity_descs = []
    data_matrix = []
    extra_rows = {}
    pce_col_values = {}
    commodity_full_totals = {}

    for row in ws.iter_rows(min_row=7, values_only=True):
        code = str(row[0]).strip() if row[0] else ''
        desc = str(row[1]).strip() if row[1] else ''

        if not code:
            continue

        # Check if this is a total/value-added row
        if code.startswith('T') or code.startswith('V'):
            vals = [safe_float(row[i]) for i in industry_indices]
            extra_rows[code] = vals
            continue

        # Regular commodity row
        commodity_codes.append(code)
        commodity_descs.append(desc)
        vals = [safe_float(row[i]) for i in industry_indices]
        data_matrix.append(vals)

        # PCE column value
        if f01000_idx is not None:
            pce_col_values[code] = safe_float(row[f01000_idx])

        # Full commodity total (intermediate + final demand, excl F05000)
        commodity_full_totals[code] = sum(safe_float(row[i]) for i in all_use_indices)

    wb.close()
    print(f'    {len(commodity_codes)} commodities x {len(industry_codes)} industries, '
          f'{len(final_demand_indices)} final demand columns')
    return (commodity_codes, commodity_descs, industry_codes,
            data_matrix, extra_rows, pce_col_values, commodity_full_totals)


def extract_use_tables():
    """Extract import use, domestic use, industry output, PCE weights,
    full commodity use totals, and industry variable cost."""

    # Read import use table
    (comm_codes, comm_descs, ind_codes,
     import_matrix, _, _, import_full_totals) = read_use_table(IMPORT_XLSX)

    # Read total use table
    (comm_codes_t, _, ind_codes_t,
     total_matrix, extra_rows, pce_values,
     total_full_totals) = read_use_table(USE_XLSX)

    # Verify alignment
    assert comm_codes == comm_codes_t, 'Commodity codes differ between import and total use tables'
    assert ind_codes == ind_codes_t, 'Industry codes differ between import and total use tables'

    # ---- bea_use_import.csv ----
    header = ['bea_code'] + ind_codes
    rows = [[comm_codes[i]] + [int(v) for v in import_matrix[i]]
            for i in range(len(comm_codes))]
    write_csv(os.path.join(OUTPUT_DIR, 'bea_use_import.csv'), header, rows)

    # ---- bea_use_domestic.csv (total - import) ----
    domestic_matrix = []
    for i in range(len(comm_codes)):
        row = [int(total_matrix[i][j] - import_matrix[i][j])
               for j in range(len(ind_codes))]
        domestic_matrix.append(row)

    rows = [[comm_codes[i]] + domestic_matrix[i]
            for i in range(len(comm_codes))]
    write_csv(os.path.join(OUTPUT_DIR, 'bea_use_domestic.csv'), header, rows)

    # ---- bea_industry_output.csv (from T008 row) ----
    if 'T008' not in extra_rows:
        raise ValueError('T008 (Total Industry Output) row not found in total use table')
    t008 = extra_rows['T008']
    rows = [[ind_codes[j], int(t008[j])] for j in range(len(ind_codes))]
    write_csv(os.path.join(OUTPUT_DIR, 'bea_industry_output.csv'),
              ['bea_code', 'output'], rows)

    # ---- bea_pce_by_commodity.csv (from F01000 column) ----
    rows = [[code, int(pce_values.get(code, 0))] for code in comm_codes]
    write_csv(os.path.join(OUTPUT_DIR, 'bea_pce_by_commodity.csv'),
              ['bea_code', 'pce'], rows)

    # ---- bea_commodity_descriptions.csv ----
    rows = [[comm_codes[i], comm_descs[i]] for i in range(len(comm_codes))]
    write_csv(os.path.join(OUTPUT_DIR, 'bea_commodity_descriptions.csv'),
              ['Commodity Code', 'Description'], rows)

    # ---- bea_commodity_use_totals.csv (full commodity absorption incl final demand) ----
    # Used for omega_M computation per Boston Fed appendix (April 2025):
    # omega_M = import_total / (import_total + domestic_total)
    rows = []
    for code in comm_codes:
        imp = import_full_totals.get(code, 0.0)
        tot = total_full_totals.get(code, 0.0)
        dom = tot - imp
        rows.append([code, int(round(imp)), int(round(dom))])
    write_csv(os.path.join(OUTPUT_DIR, 'bea_commodity_use_totals.csv'),
              ['bea_code', 'import_total', 'domestic_total'], rows)

    # ---- bea_industry_variable_cost.csv (for constant-percentage B normalization) ----
    # Per Boston Fed appendix: constant-percentage markup normalizes B matrices by
    # total intermediates + compensation of employees (not industry output)
    if 'T005' not in extra_rows:
        raise ValueError('T005 (Total Intermediate) row not found in total use table')
    v001_key = 'V00100' if 'V00100' in extra_rows else 'V001'
    if v001_key not in extra_rows:
        raise ValueError('V001/V00100 (Compensation of employees) not found')
    t005 = extra_rows['T005']
    v001 = extra_rows[v001_key]
    rows = [[ind_codes[j], int(round(t005[j] + v001[j]))]
            for j in range(len(ind_codes))]
    write_csv(os.path.join(OUTPUT_DIR, 'bea_industry_variable_cost.csv'),
              ['bea_code', 'variable_cost'], rows)

    return comm_codes


# =============================================================================
# IxC domestic requirements extraction
# =============================================================================

def extract_requirements():
    """Extract IxC domestic requirements matrix.

    Layout:
      Rows 4-5: headers (names/codes)
      Rows 6+: industry data until T010 row
      Cols 1-2: industry code + description
      Cols 3+: commodity codes
    """
    print(f'  Reading {os.path.basename(IXC_XLSX)} sheet {SHEET}...')
    wb = openpyxl.load_workbook(IXC_XLSX, read_only=True, data_only=True)
    ws = wb[SHEET]

    # Row 5 has commodity codes
    row5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    commodity_codes = [str(v).strip() for v in row5[2:] if v and str(v).strip()]

    # Read industry rows (row 6 onwards, until T010)
    industry_codes = []
    data_rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = str(row[0]).strip() if row[0] else ''
        if not code or code.startswith('T'):
            continue
        industry_codes.append(code)
        vals = [safe_float(row[i]) for i in range(2, 2 + len(commodity_codes))]
        data_rows.append(vals)

    wb.close()

    # Write bea_requirements_domestic.csv
    header = ['bea_code'] + commodity_codes
    rows = [[industry_codes[i]] + data_rows[i] for i in range(len(industry_codes))]
    write_csv(os.path.join(OUTPUT_DIR, 'bea_requirements_domestic.csv'), header, rows)

    print(f'    {len(industry_codes)} industries x {len(commodity_codes)} commodities')


# =============================================================================
# PCE bridge extraction
# =============================================================================

def extract_pce_bridge():
    """Extract PCE bridge table from PCEBridge_Detail.xlsx 2017 sheet.

    Layout:
      Row 5: headers (NIPA Line, PCE Category, Commodity Code, Commodity Descr,
              Producers' Value, Transportation, Wholesale, Retail, Purchasers' Value, Year)
      Row 6+: data
    """
    print(f'  Reading {os.path.basename(PCE_BRIDGE_XLSX)} sheet {SHEET}...')
    wb = openpyxl.load_workbook(PCE_BRIDGE_XLSX, read_only=True, data_only=True)
    ws = wb[SHEET]

    rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        nipa_line = row[0]
        pce_category = row[1]
        bea_code = row[2]
        commodity_desc = row[3]
        producers_value = safe_float(row[4])
        purchasers_value = safe_float(row[8])

        if nipa_line is None or bea_code is None:
            continue

        rows.append([
            int(nipa_line),
            str(pce_category).strip(),
            str(bea_code).strip(),
            str(commodity_desc).strip(),
            int(producers_value),
            int(purchasers_value)
        ])

    wb.close()

    write_csv(os.path.join(OUTPUT_DIR, 'pce_bridge.csv'),
              ['nipa_line', 'pce_category', 'bea_code', 'commodity_description',
               'producers_value', 'purchasers_value'],
              rows)


# =============================================================================
# GTAP-BEA crosswalk expansion
# =============================================================================

def extract_crosswalk():
    """Build detail-level GTAP-BEA crosswalk by expanding summary crosswalk.

    The summary crosswalk maps GTAP codes to summary BEA codes (e.g., PDR -> 311FT).
    The NAICS Codes sheet maps detail codes to summary codes (e.g., 311111 -> 311FT).
    We expand each GTAP -> summary mapping into GTAP -> detail mappings.

    NAICS Codes sheet layout (hierarchical, one code per row):
      Col A (Sector) -> Col B (Summary) -> Col C (U.Summary) -> Col D (Detail)
      Each level's DESCRIPTION appears in the NEXT column to the right of the CODE.
      The first non-empty column determines the hierarchy level.
    """
    # Read NAICS Codes sheet to build detail -> summary mapping
    print('  Reading NAICS Codes sheet for detail-to-summary mapping...')
    wb = openpyxl.load_workbook(USE_XLSX, read_only=True, data_only=True)
    ws = wb['NAICS Codes']

    detail_to_summary = {}
    detail_descriptions = {}
    current_summary = None
    for row in ws.iter_rows(min_row=6, max_col=8, values_only=True):
        vals = [str(v).strip() if v else '' for v in row]

        # Find the first non-empty column to determine hierarchy level
        first_col = None
        for i, v in enumerate(vals[:5]):
            if v:
                first_col = i
                break

        if first_col is None:
            continue

        if first_col == 0:
            # Sector row (col A = sector code, col B = sector description)
            continue
        elif first_col == 1:
            # Summary row (col B = summary code, col C = description)
            current_summary = vals[1]
        elif first_col == 2:
            # U.Summary row (col C = U.Summary code, col D = description)
            continue
        elif first_col == 3:
            # Detail row (col D = detail code, col E = description)
            detail_code = vals[3]
            detail_to_summary[detail_code] = current_summary
            if vals[4]:
                detail_descriptions[detail_code] = vals[4]

    wb.close()
    print(f'    {len(detail_to_summary)} detail -> summary mappings (from NAICS sheet)')

    # Filter to only detail codes present in the actual 2017 use tables.
    # The NAICS Codes sheet lists all NAICS detail codes, but the 2017 I-O
    # tables only cover ~402 commodities. Phantom codes (e.g. 331314) would
    # get tariff shocks assigned but then silently dropped during matrix
    # alignment, losing part of their parent summary code's effect.
    use_import_path = os.path.join(OUTPUT_DIR, 'bea_use_import.csv')
    with open(use_import_path, 'r') as f:
        reader = csv.DictReader(f)
        valid_commodities = {row['bea_code'] for row in reader}
    print(f'    {len(valid_commodities)} commodities in 2017 use tables')

    dropped = {c for c in detail_to_summary if c not in valid_commodities}
    if dropped:
        print(f'    Dropping {len(dropped)} detail codes not in use tables: '
              f'{sorted(dropped)[:10]}')
    detail_to_summary = {c: s for c, s in detail_to_summary.items()
                         if c in valid_commodities}
    print(f'    {len(detail_to_summary)} detail -> summary mappings (after filtering)')

    # Build summary -> [detail] lookup
    summary_to_details = {}
    for detail, summary in detail_to_summary.items():
        if summary not in summary_to_details:
            summary_to_details[summary] = []
        summary_to_details[summary].append(detail)

    # Read summary crosswalk
    summary_crosswalk = []
    with open(SUMMARY_CROSSWALK, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            summary_crosswalk.append(row)

    # Expand: for each GTAP -> summary BEA mapping, create GTAP -> detail rows
    detail_rows = []
    unmapped_summaries = set()
    for mapping in summary_crosswalk:
        gtap_code = mapping['gtap_code']
        summary_bea = mapping['bea_code']

        details = summary_to_details.get(summary_bea, [])
        if not details:
            # Summary code has no detail breakdown (services, etc.)
            unmapped_summaries.add(summary_bea)
            continue

        for detail_code in details:
            desc = detail_descriptions.get(detail_code, summary_bea)
            detail_rows.append([gtap_code, detail_code, desc])

    if unmapped_summaries:
        print(f'    Warning: {len(unmapped_summaries)} summary codes with no detail '
              f'expansion (dropped): {sorted(unmapped_summaries)[:10]}')

    write_csv(os.path.join(OUTPUT_DIR, 'gtap_bea_crosswalk.csv'),
              ['gtap_code', 'bea_code', 'bea_description'],
              detail_rows)

    # ---- bea_summary_to_detail.csv (for tau_M disaggregation) ----
    summary_detail_rows = []
    for detail_code, summary_code in sorted(detail_to_summary.items()):
        desc = detail_descriptions.get(detail_code, '')
        summary_detail_rows.append([summary_code, detail_code, desc])

    write_csv(os.path.join(OUTPUT_DIR, 'bea_summary_to_detail.csv'),
              ['summary_code', 'detail_code', 'detail_description'],
              summary_detail_rows)


# =============================================================================
# Main
# =============================================================================

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Extract Excel files from zips if needed
    make_use_zip = os.path.join(IO_DIR, 'MAKE-USE-IMPORTS (AFTER REDEFINITIONS).zip')
    requirements_zip = os.path.join(IO_DIR, 'TOTAL AND DOMESTIC REQUIREMENTS.zip')

    print('Step 1: Extracting Excel files from zips...')
    extract_from_zip(make_use_zip,
                     'ImportMatrices_After_Redefinitions_Detail.xlsx', OUTPUT_DIR)
    extract_from_zip(make_use_zip,
                     'IOUse_After_Redefinitions_PRO_Detail.xlsx', OUTPUT_DIR)
    extract_from_zip(requirements_zip,
                     'IxC_Domestic_Detail.xlsx', OUTPUT_DIR)

    print('\nStep 2: Extracting use tables (import, domestic, industry output, PCE)...')
    extract_use_tables()

    print('\nStep 3: Extracting IxC domestic requirements...')
    extract_requirements()

    print('\nStep 4: Extracting PCE bridge...')
    extract_pce_bridge()

    print('\nStep 5: Building detail-level GTAP-BEA crosswalk...')
    extract_crosswalk()

    print('\nDone! Detail CSVs written to:', OUTPUT_DIR)


if __name__ == '__main__':
    main()

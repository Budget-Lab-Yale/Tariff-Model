# =============================================================================
# extract_bea_totals.py - Extract full commodity use totals and variable cost
# =============================================================================
#
# Generates two files needed for faithful Boston Fed price model replication:
#
#   1. bea_commodity_use_totals.csv
#      Full commodity use totals (intermediate + final demand) from both the
#      Import and Total Use tables. Used to compute omega_M as the share of
#      imports in TOTAL commodity absorption, not just intermediate use.
#
#   2. bea_industry_variable_cost.csv
#      Total intermediates + compensation of employees per industry. Used as
#      the B-matrix normalization denominator under the constant-percentage
#      markup assumption (per Boston Fed appendix, April 2025).
#
# Usage:
#   python src/util/extract_bea_totals.py [--level summary|detail]
#
# =============================================================================

import csv
import os
import sys

import openpyxl


PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
IO_DIR = os.path.join(PROJECT_ROOT, 'resources', 'io')


def safe_float(val):
    """Convert cell value to float, treating None/empty/--- as 0."""
    if val is None:
        return 0.0
    s = str(val).strip()
    if s in ('', '---', '...'):
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def write_csv(path, header, rows):
    """Write CSV with given header and rows."""
    with open(path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)
    print(f'  Wrote {path} ({len(rows)} rows x {len(header)} cols)')


def extract_totals(level):
    """Extract commodity use totals and industry variable cost.

    Args:
        level: 'summary' or 'detail'
    """
    if level == 'summary':
        output_dir = IO_DIR
        import_xlsx = os.path.join(
            IO_DIR, 'BEA - Import Matrix, Before Redefinitions - Summary - 2024.xlsx')
        use_xlsx = os.path.join(
            IO_DIR, 'BEA - The Use of Commodities by Industry - Summary - 2024.xlsx')
        sheet = 'Table'
        code_row = 6
        data_start_row = 8
    elif level == 'detail':
        output_dir = os.path.join(IO_DIR, 'detail')
        import_xlsx = os.path.join(
            output_dir, 'ImportMatrices_After_Redefinitions_Detail.xlsx')
        use_xlsx = os.path.join(
            output_dir, 'IOUse_After_Redefinitions_PRO_Detail.xlsx')
        sheet = '2017'
        code_row = 6
        data_start_row = 7
    else:
        raise ValueError(f'Invalid level: {level}')

    print(f'\nExtracting {level}-level totals...')

    # ---- Read column layout from total use table ----
    print(f'  Reading column layout from {os.path.basename(use_xlsx)}...')
    wb = openpyxl.load_workbook(use_xlsx, read_only=True, data_only=True)
    ws = wb[sheet]

    row_codes = list(ws.iter_rows(
        min_row=code_row, max_row=code_row, values_only=True))[0]

    # Identify column indices by type
    t001_idx = None
    industry_indices = []
    final_demand_indices = []

    for i, val in enumerate(row_codes):
        code = str(val).strip() if val else ''
        if not code:
            continue
        if code == 'T001':
            t001_idx = i
        elif i >= 2 and not code.startswith('T') and t001_idx is None:
            # Industry column (between col 2 and T001)
            industry_indices.append(i)
        elif code.startswith('F') and not code.startswith(('F04', 'F040', 'F05', 'F050')):
            # Final demand column for domestic absorption
            # Exclude F04/F040 (exports — not domestic expenditure)
            # Exclude F05/F05000 (negative import balancing entry)
            final_demand_indices.append(i)

    if t001_idx is None:
        raise ValueError('T001 column not found')

    all_use_indices = industry_indices + final_demand_indices
    print(f'    {len(industry_indices)} industry columns, '
          f'{len(final_demand_indices)} final demand columns')

    # ---- Read commodity totals and extra rows from total use table ----
    total_commodity_totals = {}
    extra_rows = {}

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        code = str(row[0]).strip() if row[0] else ''
        if not code:
            continue

        if code.startswith('T') or code.startswith('V'):
            # Extra row (T005, V001/V00100, etc.) - read industry columns only
            vals = [safe_float(row[i]) for i in industry_indices]
            extra_rows[code] = vals
            continue

        # Regular commodity row - sum intermediate + final demand
        total = sum(safe_float(row[i]) for i in all_use_indices)
        total_commodity_totals[code] = total

    wb.close()
    print(f'    {len(total_commodity_totals)} commodity totals from total use table')

    # ---- Read commodity totals from import use table ----
    # Need to find column layout separately (may differ slightly)
    print(f'  Reading {os.path.basename(import_xlsx)}...')
    wb = openpyxl.load_workbook(import_xlsx, read_only=True, data_only=True)
    ws = wb[sheet]

    imp_row_codes = list(ws.iter_rows(
        min_row=code_row, max_row=code_row, values_only=True))[0]

    imp_t001_idx = None
    imp_industry_indices = []
    imp_final_demand_indices = []

    for i, val in enumerate(imp_row_codes):
        code = str(val).strip() if val else ''
        if not code:
            continue
        if code == 'T001':
            imp_t001_idx = i
        elif i >= 2 and not code.startswith('T') and imp_t001_idx is None:
            imp_industry_indices.append(i)
        elif code.startswith('F') and not code.startswith(('F04', 'F040', 'F05', 'F050')):
            # Same filter as total-use table: exclude exports (F04) and
            # import adjustment (F05) so omega_M uses domestic absorption
            imp_final_demand_indices.append(i)

    imp_all_use_indices = imp_industry_indices + imp_final_demand_indices
    print(f'    {len(imp_industry_indices)} industry columns, '
          f'{len(imp_final_demand_indices)} final demand columns')

    import_commodity_totals = {}
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        code = str(row[0]).strip() if row[0] else ''
        if not code or code.startswith('T') or code.startswith('V'):
            continue
        total = sum(safe_float(row[i]) for i in imp_all_use_indices)
        import_commodity_totals[code] = total

    wb.close()
    print(f'    {len(import_commodity_totals)} commodity totals from import use table')

    # ---- Write bea_commodity_use_totals.csv ----
    # Commodity codes from total use table (authoritative order)
    comm_codes = list(total_commodity_totals.keys())
    rows = []
    for code in comm_codes:
        imp_total = import_commodity_totals.get(code, 0.0)
        tot_total = total_commodity_totals[code]
        dom_total = tot_total - imp_total
        rows.append([code, int(round(imp_total)), int(round(dom_total))])

    write_csv(os.path.join(output_dir, 'bea_commodity_use_totals.csv'),
              ['bea_code', 'import_total', 'domestic_total'], rows)

    # ---- Write bea_industry_variable_cost.csv ----
    # Variable cost = total intermediates (T005) + compensation of employees (V001/V00100)
    t005_key = 'T005'
    # V-code differs: V001 (summary 2024) vs V00100 (detail 2017)
    v001_key = None
    for key in extra_rows:
        if key in ('V001', 'V00100'):
            v001_key = key
            break

    if t005_key not in extra_rows:
        raise ValueError(f'T005 (Total Intermediate) row not found in {level} use table')
    if v001_key is None:
        raise ValueError(
            f'V001/V00100 (Compensation of employees) row not found in {level} use table')

    t005 = extra_rows[t005_key]
    v001 = extra_rows[v001_key]

    # Read industry codes from the column layout
    ind_codes = []
    for i in industry_indices:
        code = str(row_codes[i]).strip() if row_codes[i] else ''
        ind_codes.append(code)

    if len(t005) != len(ind_codes) or len(v001) != len(ind_codes):
        raise ValueError(
            f'Row length mismatch: {len(ind_codes)} industries, '
            f'{len(t005)} T005 values, {len(v001)} V001 values')

    rows = []
    for j in range(len(ind_codes)):
        variable_cost = t005[j] + v001[j]
        rows.append([ind_codes[j], int(round(variable_cost))])

    write_csv(os.path.join(output_dir, 'bea_industry_variable_cost.csv'),
              ['bea_code', 'variable_cost'], rows)

    # ---- Summary stats ----
    imp_totals = [import_commodity_totals.get(c, 0) for c in comm_codes]
    tot_totals = [total_commodity_totals[c] for c in comm_codes]
    nonzero = [(i, t) for i, t in zip(imp_totals, tot_totals) if t > 0]
    if nonzero:
        omega_Ms = [i / t for i, t in nonzero]
        print(f'\n  omega_M stats (from full totals, {len(nonzero)} nonzero commodities):')
        print(f'    mean: {sum(omega_Ms)/len(omega_Ms):.4f}')
        print(f'    min:  {min(omega_Ms):.4f}')
        print(f'    max:  {max(omega_Ms):.4f}')

    print(f'\n  Variable cost stats ({len(ind_codes)} industries):')
    vc_vals = [t005[j] + v001[j] for j in range(len(ind_codes))]
    io_approx = [t005[j] + v001[j] for j in range(len(ind_codes))]
    print(f'    mean: {sum(vc_vals)/len(vc_vals):,.0f}')
    print(f'    total: {sum(vc_vals):,.0f}')


def main():
    level = 'summary'
    if '--level' in sys.argv:
        idx = sys.argv.index('--level')
        if idx + 1 < len(sys.argv):
            level = sys.argv[idx + 1]

    if level == 'both':
        extract_totals('summary')
        extract_totals('detail')
    else:
        extract_totals(level)

    print('\nDone!')


if __name__ == '__main__':
    main()
